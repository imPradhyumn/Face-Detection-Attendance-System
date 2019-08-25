#from pandas import ExcelWriter
from pandas import ExcelFile,read_excel
import datetime as dt
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

def save_attendance(names,subject):
    names=set(names)
    
    #df=read_csv('attendance_data/attendance.csv',index_col='Name')
    xls = ExcelFile('attendance_data/attendance.xlsx')
    df=read_excel(xls,subject,index_col='Name')
    
    today=dt.datetime.now()
    today=today.strftime('%d/%m/%y')
    
    student_status=['Nil' for i in range(len(df))]
    df[today]=student_status #Marking NIL initially to all student

    '''df.to_csv('attendance_data/attendance.csv')
    with ExcelWriter('attendance_data/attendance.xlsx') as writer:
        df.to_excel(writer,sheet_name=subject.lower())'''
        
    '''with ExcelWriter('attendance_data/attendance.xlsx', engine='xlsxwriter') as writer:
        writer.book=excelbook
        writer.sheets = dict((ws.title, ws) for ws in excelbook.worksheets)
        #df.to_excel(writer,sheet_name=subject.lower())
        writer.save()'''

    excelbook=load_workbook('attendance_data/attendance.xlsx')
    sheet=excelbook.get_sheet_by_name(subject.lower())
    max_col_no=sheet.max_column
    max_col_letter=get_column_letter(sheet.max_column+1)
    row_1=[]

    for i in range(max_col_no):
        row_1.append(sheet.cell(row=1,column=i+1).value)
        
    if today in row_1:
        print('Attendance already marked for today !!')
        return

    #mark attendance
    for i in names:
        df.loc[i][today]='Yes'
        
    present_status=df[today].values.tolist()
    present_status.insert(0,today)
    
    #writing to excel file
    for i in range(len(df)+1):
        #print(max_col_letter+str(i+1))  #for debugging
        c=sheet[max_col_letter+str(i+1)]
        c.value=present_status[i]
    excelbook.save('attendance_data/attendance.xlsx')
  

if __name__=='__main__':
    save_attendance(names=['Pradhyumn'],subject='mechanics') 




